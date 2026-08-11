"""
integrations/spreadsheets.py

Extracts text from an uploaded spreadsheet file (.xlsx or .csv) so it can
be injected into the conversation as context — same approach as PDFs and
Google Sheets: raw text IS the grounding, no parsing/embedding layer.
"""

from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

MAX_SPREADSHEET_CHARS = 12_000  # matches the cap used for pasted Sheet links

SUPPORTED_MIME_TYPES = {
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
}


def extract_spreadsheet_text(file_bytes: bytes, mime_type: str) -> str | None:
    if mime_type == "text/csv":
        text = _decode_csv(file_bytes)
    elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        text = _extract_xlsx(file_bytes)
    else:
        return None

    if text is None or not text.strip():
        return None

    if len(text) > MAX_SPREADSHEET_CHARS:
        text = text[:MAX_SPREADSHEET_CHARS]

    return text


def _decode_csv(file_bytes: bytes) -> str | None:
    for encoding in ("utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return None


def _extract_xlsx(file_bytes: bytes) -> str | None:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return None

    sheets_text = []
    for sheet in workbook.worksheets:
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else cell for cell in row])
        sheet_csv = buf.getvalue().strip()
        if sheet_csv:
            sheets_text.append(f"--- Sheet: {sheet.title} ---\n{sheet_csv}")

    return "\n\n".join(sheets_text) if sheets_text else None
